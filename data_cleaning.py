import pandas as pd
from openpyxl import load_workbook

# Cleaning the data


# RENAMING THE COLUMNS OF THE TABLE ---------------------------------------------------------------------


dictionary_of_columns = {
    'Origin Date': ['acquisition date',
                    'acq. date',
                    'acquisition\ndate',
                    'acquisition date(4)',
                    'initial\nacquisition\ndate(17)',
                    'investment date\n(24)',
                    'purchase date',
                    'acquisition date(39)',
                    'initial\nacquisition\ndate',
                    'date',
                    'acquisitiondate(12)',
                    'acquisition date(12)',
                    '(b)\nDate Acquired'],
    'Filing Date': [],
    'Portfolio Company': ['portfolio company',
                          'Portfolio Company (k) (o)',
                                     'Company, Geographic Location, Business Description, (Industry) and Website',

                          'Portfolio Company(1)',
                          'Portfolio Company(1) (2)',
                          'Portfolio Company(1) (2) (7)',
                          'Portfolio Company(1) (2) (7)',
                          'company (1)',
                          'issuer name',
                          'industry/company',
                          'portfolio company(1)(2)(6)',
                          'portfolio company(1)(6)',
                          'portfolio company (1)(6)',
                          'portfolio company(6)',
                          'portfolio company(5)',
                          'company(1)(4)(8)(32)',
                          'company(1)(2)(3)(20)(29)',
                          'company(1)(7)(18)(20)',
                          'company(1)(2)(19)(23)',
                          'investments—non-controlled/non-affiliated(1)',
                          'investments—non-controlled/affiliated',
                          'investments—controlled/affiliated',
                          'portfolio company(a)',
                          'country/security/industry/company',
                          'portfolio company (1)(3)',
                          'company(1)',
                          'portfolio company (1) (20)',
                          'investments-non-controlled/non-affiliated(1) (2)',
                          'portfolio company, location and industry(1)',
                          'portfolio company, location and industry (1)',
                          'company(1)(9)',
                          'portfolio company 1 2 3 4',
                          'portfolio company1 2 3',
                          'investments-\nnon-controlled/\nnon-affiliated (1)',
                          'issuer',
                          'company(1)(2)',
                          'company #+',
                          'Investments-non-controlled/non-affiliated(1)(2)'],
    'Non-Accrual': [],
    'Reported Industry': ['industry',
                 'Business Description',
                 'business description',
                 'industry(2)',
                 'sector',
                 'industry(46)',
                 'description',
                 ''],
    'Reported Investment Type': ['type of investment (1)(2)(3)(4)',
                                 '(a)\nType of Investment',
                        'Investment(2)',
                        'investment type',
                        'investment',
                        'investment\ntype',
                        'investment type(1)(2)',
                        'investments(1)(19)',
                        'type of investment (7)',
                        'investment(1)(5)',
                        'type of\ninvestment',
                        'series(3)',
                        'type of investment (2) (3) (15)',
                        'investments(1)(2)(3)',
                        'type of investment',
                        'investment (2), (4), (12), (14), (23), (24)',
                        'investments(1)(35)',
                        'investments',
                        'security(2)',
                        'investment',
                        'type of warrant',
                        'type of equity',
                        'investments(1)(13)(17)',
                        'type of investment (1)(2)(3)',
                        'Investment(3)'],
    'Reference Rate': ['index',
                       'reference (7)',
                       'Reference(4)',
                       'reference rate and\nspread (4)',
                       'reference\nrate and\nspread(4)',
                       'reference\nrate\nspread (4)',
                       'reference (6)',
                       'index(1)',
                       'reference rate and spread(2)',
                       'reference rate (2)',
                       'reference rate(2)',
                       'reference rate\nand spread(3)',
                       'reference rate and spread (28)',
                       'reference rate and spread',
                       'reference rate and spread(4)',
                       'reference',
                       'reference (10)',
                       'reference rate\nand spread (1)',
                       'ref',
                       'reference\nrate and\nspread',
                       'referencerate and spread(5)',
                       'reference rate\nand spread',
                        'reference rate and spread'],
    'Spread': ['spread',
                'spreadaboveindex(1)',
                'spread (3)',
                'basis point spread above index(4)',
                'spread (2)',
                'spread(2)',
                'margin',
                'spread (10)',
                'spread\nabove\nindex(7)',
                'and spread',
                'spread (6)',
                'Spread Above Reference Rate(3)'],
    'PIK': ['pik', 'pik rate (19)', 'pik(10)', 'pik', 'PIK Rate'],
    'Stated Coupon': ['coupon (3)',
                      'Investment Coupon Rate/ Maturity (i)',
                      'Basis Point\n Spread\n Above\n Index(1)',
                      'Basis Point Spread Above Index(1)',
                      'interest rate(6)',
                      'interest\nrate',
                      'interest rate(12)',
                      'cash interest rate (5)',
                      'interest rate',
                      'interest',
                      'interest rate(2)(15)',
                      'interest rate (2)',
                      'interest rate(2)',
                      'all-in rate',
                      'cash rate (4)',
                      'interest\nterm *',
                      'rate(b)',
                      'interest rate and floor(1)',
                      'interest\nrate(3)',
                      'total rate',
                      'interest rate(3)',
                      'interest rate(5)',
                      'total coupon (17)',
                      'interest rate (10)',
                      'interest rate (1)',
                      'coupon/yield',
                      'interest\nrate(1)',
                      'coupon',
                      'rate',
                      'interest\nrate *',
                      'interest rate(2) (12)',
                      'interestrate(2)',
                      'interest rate (6)',
                      'cash interest rate (4)(5)'],
    'Calculated Yield': ['Current\n Coupon', 'current coupon', 'Current Coupon'],
    'Floor': ['floor', 'floor(b)', 'interest rate floor', 'floor (1)', ''],
    'Ceiling': ['ceiling', '', ''],
    'ETP': ['etp (10)'],
    'Maturity': ['maturity date',
                 'maturity\ndate',
                 'maturity',
                 'maturity/\ndissolution\ndate',
                 'maturity/expiration\ndate',
                 'maturity/expiration date',
                 'legal maturity',
                 'maturity5'],
    'Shares': ['shares',
               'shares/units',
               'par/shares(2)',
               'units / shares',
               'units/shares',
               'shares/ units',
               'number of\nshares',
               'shares(4)',
               'shares/\nunits',
               'shares(3)',
               'units',
               '/ units',
               'shares / units'],
    'Par': ['principal (7)',
            'Principal/ Numbers of Shares',
            '(c)\nEquity',
            'Par Amount/ Shares',
            'Par /\n Shares',
            'Par Amount/ Shares(16)',
            'Principal /\nPar',
            'principal ($) /shares(3)',
            'principal',
            'par / shares',
            'par\namount/\nshares',
            'par amount/\nshares',
            'par/shares(3)',
            'principal/shares(9)',
            'principal/\nshares(9)',
            'principal amount',
            'par amount/units(1)',
            'par/ principal amount **',
            'principal\namount(c)',
            'par amount/ llc interest',
            'principal\namount,\npar value\nor shares **',
            'par(4)',
            'principal\namount(c)/ shares',
            'principal\namount',
            'par / units',
            'par/units',
            'par amount/units',
            'principal amount(c)/shares',
            'principal (4)',
            'par amount/ shares(4)',
            'par amount / shares(6)',
            'principal\namount,\npar value\nor shares (15)',
            'principal\namount,\npar value\nor shares',
            'principal\namount,\npar value or shares',
            'par/shares (++)',
            'principal value',
            'par amount /\nshares',
            'par amount / shares',
            'par amount',
            'par amount/ units',
            'par\namount',
            'par\namount/shares',
            'principal/ par amount(3)',
            'outstanding\nprincipal',
            'paramount/units(1)',
            'principal (6)'],
    'Cost': ['cost',
             'Cost(2)(3)',
             'Cost(3)(4)',
             'Cost(4)',
             'amortized cost',
             'cost(37)',
             'amortized\ncost(5)',
             'cost(28)',
             'cost(3)',
             'amortized cost(2)(3)',
             'amortized cost(4)(25)',
             'amortized cost(4)(5)',
             'amortized cost(3)(4)',
             'amortized cost (4)',
             'amortized cost(4)',
             'amortized\ncost',
             'amortized\ncost',
             'cost(2)',
             'cost of investments (6)(9)',
             'cost of',
             'amortized cost(3)',
             'cost (4)',
             'cost(5)',
             'cost(7)',
             'cost',
             'amortized\ncost(2) (7)',
             'amortized\ncost(2)(7)',
             'cost6',
             'cost 4',
             'cost (3)',
             'cost(6)',
             'amortized\ncost(++)',
             'amortizedcost(3)'],
    'FV': ['fair value',
           '(d)(f)\nFair Value',
           'Fair Value(4)',
           'Fair Value(2)',
           'fairvalue(4)',
           'fair value (3)',
           'fair value(1)(38)',
           'fair\nvalue',
           'fair value(1)(29)',
           'market value',
           'market\nvalue',
           'fair value (5)',
           'fair value(5)',
           'fairvalue(5)',
           'fair\nvalue (5)',
           'fairvalue(d)',
           'fair\nvalue(d)',
           'fair value (9)',
           'value',
           'fair',
           'fair value(d)',
           'fair value (18)',
           'fair\nvalue',
           'fair\nvalue(2)',
           'fair value(6)',
           'fair value (8)',
           'value(1)'],
    'Notes': ['notes', 'footnotes','Footnotes(1)(2)'],
    'Cash': ['cash'],
    'Assets': ['assets7', 'assets 5']
}


# Rename for missing columns
bdc_missing_columns = ["Main Street Capital CORP",
                       "ARES CAPITAL CORP", "GOLUB CAPITAL BDC, Inc."]


def function_to_add_columns_for_unnamed_columns(df, bdc_name):

    if bdc_name == "GOLUB CAPITAL BDC, Inc.":
        df.rename(columns={"None": "Portfolio Company"}, inplace=True)
    else:
        df.rename(columns={"None": "Notes"}, inplace=True)

    return df


# Rename columns based on the dictionary


def function_to_rename_columns(df):
    rename_mapping = {}
    # Ensure all column names are strings
    df.columns = df.columns.astype(str)


    for key, values in dictionary_of_columns.items():
        # Normalize values to be lists
        if not isinstance(values, (list, tuple)):
            values = [values]

        # Clean and normalize possible values for better matching
        values = [value.strip().lower()
                  for value in values if isinstance(value, str)]

        for col in df.columns:
            col_cleaned = col.strip().lower()  # Normalize column name for matching

            # Use exact matching instead of substring matching
            if col_cleaned in values:
                rename_mapping[col] = key

    # Apply the renaming
    df.rename(columns=rename_mapping, inplace=True)
    return df

# ---------------------------------------------------------------------------------------------------


# Data Cleaning New --------------------------------------------------------------------------------------


# def adjust_sheets_main(sheet, bdc_name):
#     # Step 1: Identify merged column ranges in the first row
#     merged_ranges = []
#     merged_columns = set()
    

#     for merged_range in sheet.merged_cells.ranges:
#         min_col, min_row, max_col, max_row = merged_range.bounds
#         if min_row == 1:  # If merged in the first row
#             header_value = sheet.cell(row=min_row, column=min_col).value
#             merged_ranges.append((header_value, min_col, max_col))
#             # Track merged columns
#             merged_columns.update(range(min_col, max_col + 1))

#     # Step 2: Identify non-merged columns
#     max_col = sheet.max_column  # Get total number of columns
#     non_merged_columns = [
#         col for col in range(1, max_col + 1) if col not in merged_columns
#     ]

#     # Step 3: Process merged columns
#     result = {}
#     max_row = sheet.max_row  # Get last row with data

#     for header, min_col, max_col in merged_ranges:
#         concatenated_data = []

#         for row in range(2, max_row + 1):
#             combined_row = [
#                 str(sheet.cell(row=row, column=col).value) if sheet.cell(
#                     row=row, column=col).value is not None else ""
#                 for col in range(min_col, max_col + 1)
#             ]
#             concatenated_data.append(" ".join(combined_row))

#         result[header] = concatenated_data

#     # Step 4: Process non-merged columns
#     for col in non_merged_columns:
#         header = sheet.cell(row=1, column=col).value  # Get column header
#         column_data = [
#             sheet.cell(row=row, column=col).value if sheet.cell(
#                 row=row, column=col).value is not None else ""
#             for row in range(2, max_row + 1)
#         ]
#         result[header] = column_data

#     # Step 5: Convert dictionary to DataFrame
#     unnamed = pd.DataFrame(result)
#     # print(unnamed.columns)
#     # df = function_to_rename_columns(unnamed)  # Apply renaming function
#     # if bdc_name in bdc_missing_columns:
#     #     function_to_add_columns_for_unnamed_columns(df, bdc_name)

#     # print(df.columns)
#     # print(df["None"])
#     return unnamed  



def adjust_sheets_main(sheet, bdc_name):
    # Step 1: Identify merged column ranges in the first row
    merged_ranges = []
    merged_columns = set()

    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_row == 1:  # If merged in the first row
            header_value = sheet.cell(row=min_row, column=min_col).value
            if header_value is None:
                header_value = f"Unnamed_{min_col}"
            merged_ranges.append((header_value, min_col, max_col))
            merged_columns.update(range(min_col, max_col + 1))

    # Step 2: Identify non-merged columns
    max_col = sheet.max_column
    non_merged_columns = [
        col for col in range(1, max_col + 1) if col not in merged_columns
    ]

    # Step 3: Process merged columns
    result = {}
    max_row = sheet.max_row

    for header, min_col, max_col in merged_ranges:
        concatenated_data = []
        for row in range(2, max_row + 1):
            combined_row = [
                str(sheet.cell(row=row, column=col).value) if sheet.cell(
                    row=row, column=col).value is not None else ""
                for col in range(min_col, max_col + 1)
            ]
            concatenated_data.append(" ".join(combined_row).strip())
        result[header] = concatenated_data

    # Step 4: Process non-merged columns
    for col in non_merged_columns:
        header = sheet.cell(row=1, column=col).value
        if header is None:
            header = f"Unnamed_{col}"
        column_data = [
            sheet.cell(row=row, column=col).value if sheet.cell(
                row=row, column=col).value is not None else ""
            for row in range(2, max_row + 1)
        ]
        result[header] = column_data

    # Step 5: Convert dictionary to DataFrame
    unnamed = pd.DataFrame(result)
    return unnamed

# Post Processing ----------------------------------------------------------------------------------------------

# Biforcating Industry, Investment and Portfolio company


def seperate_investments_portfolio_industry(original_Dataframe):
    df = pd.read_excel("./SOI-AI database.xlsx", sheet_name="refined-Unique")
    Industries_list = list(df["Industry Unique"])
    Investment_type = list(df["Investment Unique"])
    count = 0
    # Initialize the new DataFrame
    newDf = pd.DataFrame(
        columns=["Industry", "Investment", "Portfolio Company"])

    # Ensure all items in Industries_list and Investment_type are strings
    Industries_list = [str(industry) for industry in Industries_list]
    Investment_type = [str(investment) for investment in Investment_type]

    #
    for i in range(len(original_Dataframe)):
        word = original_Dataframe["Investment"].iloc[i]
        # Ensure the word is a valid string, strip extra spaces, and check for non-empty strings
        if isinstance(word, str):
            word = word.strip()

            if word:  # Ensure the word is not empty
                # Corrected condition to check if `word` is a substring in any list item
                if (word == "nan"):
                    newDf.loc[i, "Industry"] = ""
                elif any(word in substring for substring in Industries_list):
                    newDf.loc[i, "Industry"] = word
                elif any(word in substring for substring in Investment_type):
                    newDf.loc[i, "Investment"] = word
                else:
                    newDf.loc[i, "Portfolio Company"] = word
            else:
                # Handle empty strings or strings with only spaces
                newDf.loc[i, "Portfolio Company"] = word
        count += 1
    newDf[["Industry", "Investment"]] = newDf[[
        "Industry", "Investment"]].ffill()
    # original_Dataframe = original_Dataframe["Investment"] = ""
    # final_Df = pd.concat([original_Dataframe, newDf])
    original_Dataframe[["Industry", "Investment", "Portfolio Company"]] = newDf[[
        "Industry", "Investment", "Portfolio Company"]]
    return original_Dataframe


