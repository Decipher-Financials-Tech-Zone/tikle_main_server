import requests
import tempfile
from bs4 import BeautifulSoup
import re
from fastapi import FastAPI, Query, HTTPException, File, UploadFile, Form, Request, Response
import httpx
from pydantic import BaseModel, HttpUrl
from io import BytesIO
import io
import os
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import numpy as np
from data_cleaning import *
from authentication import *
from typing import Optional, Dict, List, Final
from new_docs_notification import *
from datetime import datetime, date
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import numbers
from filter_portfolio_functions import find_portfolio_end, find_portfolio_start
from models.bdc_status import Bdc_status
from config.database import collection_name
from schema.schemas import get_all_bdc_status, get_one_bdc_status
from bson import ObjectId
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse
from integrated_ai.claude_portfolio_batch_processing import *
import json
from clean_portfolio import *
from miscellaneous_functions import *
from financial_statements import extract_first_table_from_sec, fetch_and_parse_filing_summary, group_reports_by_category
from agentic_ai_toolkit.agentic_ai_main import append_excel_tabs, process_header_renaming, process_bdc, process_excel_file, extract_first_of_contiguous_group
import fitz  # PyMuPDF
import zipfile
from daily_task_sheet import fetch_tasks_for_today, send_email, build_html_table_grouped
import asyncio
import xml.etree.ElementTree as ET
from fastapi.encoders import jsonable_encoder
from urllib.parse import unquote
import json
from pathlib import Path
from fastapi.concurrency import run_in_threadpool
import logging

# Build path to cases.json inside the package
json_path = Path("agentic_ai_toolkit") / "cases.json"

with open(json_path, "r", encoding="utf-8") as f:
    cases = json.load(f)


# Start the FastAPI application

app = FastAPI()


HEADERS = {"user-agent": "luv.ratan@decipherfinancials.com"}
# Define the allowed origins (your React/Next.js app's URL)
origins = [
    "http://localhost:3000",  
    "http://localhost:3010",
    "http://13.201.129.153:3000",
    "http://13.201.129.153:3010",

    # Old Netlify test deployments
    "https://tickledecipherqc.netlify.app",
    "https://tickledecipher.netlify.app",
    "https://tickleadmin.netlify.app",
    "https://tikle.netlify.app",

    # Render test deployments
    "https://new-main-application-frontend.onrender.com",
    "https://new-main-application-frontend.onrender.com/dashboard",
    "https://new-main-application-frontend.onrender.com/dashboard/view-filings",

    # 🚀 Production domain
    "https://tikle.in"
]

# Add CORS middleware to FastAPI
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,        # Allow these specific origins
    allow_credentials=True,
    allow_methods=["*"],          # Allow all HTTP methods
    allow_headers=["*"],          # Allow all headers
)


# FUNCTION TO FETCH DATA --------------------------------------------------------------------------------------


def function_to_adjust_url(given_url):
    base_url = "https://www.sec.gov"
    strArr = given_url.split("=")
    finalUrl = base_url + strArr[1]
    return finalUrl


def sec_to_excel(url: str, bdc_name: str):
    headers = {"User-Agent": "luv.ratan@decipherfinancials.co.in"}

    response = requests.get(url, headers=headers)
    if response.status_code != 200:
        raise HTTPException(
            status_code=response.status_code, detail="Failed to fetch the URL"
        )

    soup = BeautifulSoup(response.content, "html.parser")

    # Extract tables
    tables = soup.find_all("table")
    portfolio_tables_array = []
    count = 0

    print(bdc_name)

    for table in tables:
        if bdc_name == "GUGGENHEIM CREDIT INCOME FUND":
            if (
                "xdx: Statement - CONSOLIDATED SCHEDULE OF INVESTMENTS (UNAUDITED)" in str(table)
                or "us-gaap:InvestmentOwnedBalancePrincipalAmount" in str(table)
            ):
                portfolio_tables_array.append(count)

        elif (
            "InvestmentOwnedAtCost" in str(table)
            and "InvestmentOwnedAtFairValue" in str(table)
            and "us-gaap:Liabilities" not in str(table)
        ):
            if find_portfolio_start(bdc_name, str(table)):
                count += 1
                continue

            portfolio_tables_array.append(count)

            if find_portfolio_end(bdc_name, str(table)):
                break
        count += 1

    final_porfolio_list = portfolio_tables_array[:]
    if bdc_name not in [
        "FS KKR Capital Corp",
        "KKR FS Income Trust",
        "FS Specialty Lending Fund",
        "Golub Capital Direct Lending Unlevered Corp",
    ]:
        for i in range(1, len(portfolio_tables_array)):
            if portfolio_tables_array[i] != portfolio_tables_array[i - 1] + 1:
                final_porfolio_list = portfolio_tables_array[:i]
                break

    # Build structured list of tables
    table_dicts = [
        {"Table_Index": idx, "Table_HTML": str(tables[idx])}
        for idx in final_porfolio_list
    ]

    # ✅ Return only tables, no full HTML
    return {"tables": table_dicts}


def all_tables(url: str, bdc_name: str):
    headers = {"User-Agent": "luv.ratan@decipherfinancials.co.in"}

    adjusted_url = url

    response = requests.get(adjusted_url, headers=headers)

    if response.status_code != 200:
        raise HTTPException(
            status_code=response.status_code, detail="Failed to fetch the URL"
        )

    soup = BeautifulSoup(response.content, "html.parser")

    # New Code starts here---------------------------------------------
    # Extract tables
    tables = soup.find_all("table")
    portfolio_tables_array = []

    count = 0

    for table in tables:

        portfolio_tables_array.append(count)
        count += 1

    final_porfolio_list = portfolio_tables_array[:]  # Default to full list

    table_dicts = []
    for items in final_porfolio_list:
        table_dict = {
            "Table_Index": items,
            # Store the HTML content of the table
            "Table_HTML": str(tables[items]),
        }
        table_dicts.append(table_dict)

    return {
        "full_html": "str(soup)",  # Full HTML as a string
        "tables": table_dicts,  # List of extracted tables
    }


# FUNCTION TO CREATE EXCEL SHEET FROM THE SELECTED TABLES -----------------------------------------------------------

required_rows_array = [
    "Company Name",
    "Ticker",
    "Filing Date",
    "DF Co. ID",
    "DF Asset ID",
    "DF Sec ID",
    "Portfolio Company",
    "Unified name",
    "Non-Accrual",
    "Reported Industry",
    "DF Industry",
    "Reported Investment Type",
    "DF Investment Type",
    "Reference Rate",
    "Spread",
    "PIK",
    "Stated Coupon",
    "Calculated Yield",
    "Floor",
    "Ceiling",
    "ETP",
    "Origin Date",
    "Maturity",
    "Shares",
    "Par",
    "Cost",
    "FV",
    "$ Mark",
    "% Mark",
    "Notes",
]
required_rows_df = pd.DataFrame(columns=required_rows_array)

# arr = ['BDC Name', 'Portfolio Company','Reported Industry', 'Reported Investment Type', 'Reference Rate',
#         'Spread',"PIK", 'Stated Coupon', 'Calculated Yield', 'Floor', 'Ceiling','ETP', 'Origin Date', 'Maturity',
#         'Shares', 'Par', 'Cost', 'FV', 'Notes']


def function_to_merge_sheets(wb, filing_date, bdc_name):
    all_dataframes = []
    all_column_heads = []

    # Iterate over each sheet

    for sheet_name in wb.sheetnames:

        sheet = wb[sheet_name]
        print(f"Processing sheet: {sheet_name}")

        # Apply checkFunction (adjust_sheets_main in this case) on the sheet
        processed_df = adjust_sheets_main(sheet, bdc_name)
        # llm_processing = function_to_manipulate_table_through_claude(processed_df)
        # all_column_heads.append(processed_df.columns.to_list())

        # Append to the list of DataFrames\
        # For direct AI
        data = sheet.values
        columns = next(data)  # Get the first row as header
        df = pd.DataFrame(data, columns=columns)
        all_dataframes.append(processed_df)

    return all_dataframes


def format_number_columns(ws):
    """
    Formats specified columns as numbers, percentages, or dates.

    Args:
        ws (Worksheet): OpenPyXL worksheet.
        percent_cols (list): List of column letters to format as percentages.
    """
    number_cols = ["Y", "Z", "AA"]
    percent_cols = ["N", "O", "P", "Q", "R", "S", "T", "U"]  # J column as percentage
    date_cols = ["A", "B", "M"]  # Example date columns

    for col_letter in number_cols + percent_cols + date_cols:
        # Convert letter to column index
        col_idx = column_index_from_string(col_letter)

        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    try:
                        cell.value = float(cell.value)  # Convert to float

                        # Apply formatting
                        if col_letter in number_cols:
                            cell.number_format = "0.00"  # Two decimal places
                        elif col_letter in percent_cols:
                            cell.number_format = (
                                "0.00%"  # Percentage format with two decimal places
                            )

                    except ValueError:
                        pass  # Ignore if conversion fails


# FUNCTION TO FETCH ACCESSION NUMBERS FOR ALL THE FILES -----------------------------------------------------------


def function_to_adjust_meta_Data(data):
    filings = data["filings"]
    recent = pd.DataFrame(filings["recent"])

    # # Filter rows where 'primaryDocDescription' is "10-K", "10-Q", or "8-K"
    # recent_filtered = recent[
    #     (recent["form"] == "10-K") |
    #     (recent["form"] == "10-Q") |
    #     (recent["form"] == "8-K")
    # ]

    # Convert the filtered DataFrame to JSON
    recent_ten_and_eights = recent.to_json(orient="records")

    meta_data = {
        "name": data["name"],
        "cik": data["cik"],
        "fiscalYearEnd": data["fiscalYearEnd"],
        "exchanges": data["exchanges"],
        "tickers": data["tickers"],
        "recent_filings_array": recent_ten_and_eights,
    }

    return meta_data


async def get_SEC_metadata(CIK: str):
    """
    Fetches metadata for a given CIK from the SEC API.

    Parameters:
    - CIK (str): Central Index Key of the company.

    Returns:
    - JSON object with the metadata.
    """
    try:
        url = f"https://data.sec.gov/submissions/CIK{CIK}.json"
        # Replace with your email for SEC compliance
        headers = {"User-Agent": "luv.ratan@decipherfinancials.com"}

        # Fetch the data asynchronously
        async with httpx.AsyncClient() as client:
            response = await client.get(url, headers=headers)

        # Raise an error if the response status is not 200
        if response.status_code != 200:
            raise HTTPException(
                status_code=response.status_code,
                detail=f"Error fetching data: {response.text}",
            )

        # Validate and return the JSON data
        try:
            data = response.json()
        except ValueError:
            raise HTTPException(
                status_code=500, detail="Invalid JSON received from the SEC API"
            )

        final_data = function_to_adjust_meta_Data(data)

        return final_data

    except HTTPException as e:
        # Re-raise HTTP exceptions for proper FastAPI error responses
        raise e
    except Exception as e:
        # Handle unexpected errors
        raise HTTPException(status_code=500, detail=f"Unexpected error: {str(e)}")


# API ENDPOINTS -----------------------------------------------------------------------------------------------


# Fetch SEC data


@app.get("/fetch-free-sec-data")
async def get_document_data(url: str = Query(...), bdc_name: str = Query(...)):
    return sec_to_excel(url, bdc_name)


@app.get("/get-all-tables")
async def get_all_tables(url: str = Query(...), bdc_name: str = Query(...)):
    return all_tables(url, bdc_name)


# Create/Initiate Portfolio through claude


@app.post("/upload/")
async def upload_file(
    file: UploadFile = File(...),  # To handle the uploaded file
    bdc_name: str = Form(...),  # To handle the bdc_name from the form
    # To handle the filing_date from the form
    filingDate: str = Form(...),
):
    # Read the uploaded file into memory
    file_contents = await file.read()

    # Load the uploaded file as a workbook
    excel_data = BytesIO(file_contents)
    wb = load_workbook(excel_data)

    # Process all sheets and create a concatenated DataFrame
    final_concatenated_df = function_to_merge_sheets(wb, filingDate, bdc_name)
    # print(final_concatenated_df)
    response = claude_batch_processing_portfolio(final_concatenated_df)
    # print(response)
    # print(type(response))
    response_json_string = response.to_json()
    in_process_claude_response = json.loads(response_json_string)
    print(in_process_claude_response)

    bdc_status = {
        "bdc_name": bdc_name,
        "complete_file_name": f"{bdc_name} {filingDate}",
        "reporting_date": filingDate,
        "claude_batch_reference_id": in_process_claude_response["id"],
        "claude_status_of_requests": in_process_claude_response["processing_status"],
        "postgresql_status": "NA",
        "postgresql_reference_id": "NA",
    }

    doc = dict(bdc_status)
    result = collection_name.insert_one(doc)
    doc["_id"] = str(result.inserted_id)  # Add the ID to the doc for return
    return JSONResponse(
        status_code=200,
        content={"message": "BDC status added successfully", "inserted_document": doc},
    )

    # # Convert date columns to datetime format
    # date_cols = ['Maturity', 'Origin Date', 'Filing Date']
    # for col in date_cols:
    #     final_concatenated_df[col] = pd.to_datetime(final_concatenated_df[col], errors='coerce')

    # # Format as 'YYYY/MM/DD', handling NaT values safely
    # for col in date_cols:
    #     final_concatenated_df[col] = final_concatenated_df[col].apply(
    #         lambda x: x.strftime('%d/%m/%Y') if pd.notna(x) else ''
    #     )

    # # Load the template Excel file
    # template_path = "./soi_output.xlsx"  # Path to your template file
    # template_wb = load_workbook(template_path)
    # template_ws = template_wb["template"]

    # # Add Post formatting here----------------------------------------------------------------------------
    # # bdc_list_for_separation = ["Blackstone Secured Lending Fund", "Blackstone Private Credit Fund"]

    # # if bdc_name in bdc_list_for_separation:
    # #     final_concatenated_df = seperate_investments_portfolio_industry(final_concatenated_df)
    # # # Post formatting ends here ---------------------------------------------------------------------------

    # # Write DataFrame contents starting from row 2
    # start_row = 2
    # for r_idx, row in enumerate(dataframe_to_rows(final_concatenated_df, index=False, header=True), start=start_row):
    #     for c_idx, value in enumerate(row, start=1):
    #         cell = template_ws.cell(row=r_idx, column=c_idx, value=value)

    # # Auto-adjust the width of column C based on its content --------------------------------------------------------
    # template_ws.column_dimensions["G"].width = 50
    # template_ws.column_dimensions["G"].width = 30
    # format_number_columns(template_ws)
    # template_ws.freeze_panes = "A3"

    # # Save the modified template to memory
    # output_buffer = BytesIO()
    # template_wb.save(output_buffer)
    # output_buffer.seek(0)

    # return StreamingResponse(
    #     open("test.xlsx", "rb"),
    #     media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    #     headers={"Content-Disposition": "attachment; filename=test.xlsx"}
    # )


# Fetch portfolio data from claude and format it.


@app.post("/fetch_clean_portfolio")
async def get_clean_portfolio(
    claude_reference_id: str = Form(...),
    bdc_name: str = Form(...),
    filingDate: str = Form(...),
):
    print(claude_reference_id, bdc_name, filingDate)

    df = fetch_portfolio_data_from_claude(claude_reference_id)

    # template_path = "./soi_output.xlsx"  # Path to your template file
    # template_wb = load_workbook(template_path)
    # template_ws = template_wb["template"]

    # # Write DataFrame contents starting from row 2
    # start_row = 2
    # for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
    #     for c_idx, value in enumerate(row, start=1):
    #         cell = template_ws.cell(row=r_idx, column=c_idx, value=value)

    # # Auto-adjust the width of column C based on its content --------------------------------------------------------
    # template_ws.column_dimensions["G"].width = 50
    # template_ws["B1"] = "BDC Name here"
    # format_number_columns(template_ws)
    # template_ws.freeze_panes = "A3"

    # # Save the modified template to memory
    # output_buffer = BytesIO()
    # template_wb.save(output_buffer)
    # output_buffer.seek(0)

    df.to_excel("Portfolio.xlsx", engine="openpyxl", index=None)

    return StreamingResponse(
        open("Portfolio.xlsx", "rb"),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=test.xlsx"},
    )

    # Convert DataFrame to a list of records (dicts)
    # return JSONResponse(content={"data": df.to_dict(orient="records")})


@app.get("/get_BDC_Meta_File")
async def get_mets_Data(cik: str = Query(...)):
    return await get_SEC_metadata(cik)


@app.get("/")
def homePage():
    return "TIKLE main backend Home page"


# Login endpoint
@app.post("/login")
async def login(user: LoginRequest):
    # Check if the user exists
    db_user = await users_collection.find_one({"email": user.email})
    if not db_user:
        raise HTTPException(status_code=400, detail="Invalid credentials")

    # Verify the password
    if not pwd_context.verify(user.password, db_user["password"]):
        raise HTTPException(status_code=400, detail="Invalid credentials")

    # Generate JWT token
    token_data = {"sub": db_user["email"]}
    access_token = create_access_token(data=token_data)
    return {"authToken": access_token}


# Validate user endpoint
@app.get("/validate-user")
async def validate_user(current_user: dict = Depends(get_current_user)):
    return {
        "isValid": True,
        "user": {
            "id": str(current_user["_id"]),
            "name": current_user["name"],
            "email": current_user["email"],
        },
    }


# Fetch the latest documents notification from the database


@app.get("/documents/")
async def get_documents():
    """Fetch all documents from the collection."""
    # documents = await collection.find().to_list(length=100)  # Limit to 100 docs
    documents = await collection.find().to_list()
    # Convert ObjectId to string for proper JSON serialization
    for doc in documents:
        doc["_id"] = str(doc["_id"])
    return documents


# Database functions ----------------------------------------------------------------------------------------------------------


def add_new_bdc_status_function(bdc_status: Bdc_status):
    doc = dict(bdc_status)
    result = collection_name.insert_one(doc)
    doc["_id"] = str(result.inserted_id)  # Add the ID to the doc for return
    return {"message": "BDC status added successfully", "inserted_document": doc}
    # return JSONResponse(
    #     status_code=200,
    #     content={
    #         "message": "BDC status added successfully",
    #         "inserted_document": doc
    #     }
    # )


# Fetch the portfolio status of all the BDC


@app.get("/all_bdc_status")
async def all_bdc_status():
    status_list = get_all_bdc_status(collection_name.find())
    return status_list


# Fetch the Portfolio Status of the given bdc by file name


@app.get("/bdc_status_by_file/{complete_file_name}")
async def get_bdc_status_by_file(complete_file_name: str):
    # Find the document in the collection
    bdc_status = collection_name.find_one({"complete_file_name": complete_file_name})

    if not bdc_status:
        raise HTTPException(status_code=404, detail="Document not found")

    # Check if processing is in progress
    if bdc_status["claude_status_of_requests"] == "in_progress":
        try:
            # Fetch the latest status from Claude
            batch_id = bdc_status["claude_batch_reference_id"]
            response = fetch_status_from_claude(batch_id)

            # Check if processing has ended
            if response.processing_status == "ended":
                # Update the status in the database
                bdc_status["claude_status_of_requests"] = "ended"

                # Update the document in the collection
                result = collection_name.update_one(
                    {"complete_file_name": complete_file_name},
                    {"$set": bdc_status},  # Make sure this works with your data model
                )

                if result.matched_count == 0:
                    raise HTTPException(
                        status_code=404, detail="Failed to update document"
                    )

                return {"message": "Document updated successfully", "status": "ended"}

            # If not ended, return the current status
            return get_one_bdc_status(bdc_status)

        except Exception as e:
            # Handle any unexpected errors
            raise HTTPException(
                status_code=500, detail=f"Error processing request: {str(e)}"
            )

    # If not in progress, return the current status
    return get_one_bdc_status(bdc_status)


@app.post("/add_bdc_status")
async def add_new_bdc_status(bdc_status: Bdc_status):
    return add_new_bdc_status_function(bdc_status)


@app.put("/{complete_file_name}")
async def update_status_of_file(complete_file_name: str, bdc_status: Bdc_status):
    result = collection_name.update_one(
        {"complete_file_name": complete_file_name}, {"$set": bdc_status.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Document not found")
    return {"message": "Document updated successfully"}


# function to import Financial Statements


@app.get("/sec-filing/tables-with-html")
async def get_sec_tables_with_html(
    cik: str = Query(..., description="CIK number of the company"),
    accession_number: str = Query(..., description="Accession number without dashes"),
):
    # Construct URLs
    accession_number_nodash = accession_number.replace("-", "")
    sec_html_url = f"https://www.sec.gov/cgi-bin/viewer?action=view&cik={cik}&accession_number={accession_number}"
    base_table_url = (
        f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{accession_number_nodash}/"
    )

    HEADERS = {"User-Agent": "luv.ratan@decipherfinancials.co.in"}

    async with httpx.AsyncClient() as client:
        r = await client.get(sec_html_url, headers=HEADERS)

        if r.status_code == 200:
            links = extract_links_from_menu_cat3(r.text)
            results = []

            for link in links:
                id_ = link["id"]
                url = f"{base_table_url}{id_.upper()}.htm"
                table_html = await extract_first_table(url)
                results.append(
                    {"id": id_, "text": link["text"], "table_html": table_html}
                )

            return JSONResponse(content={"data": results})

        return JSONResponse(
            content={"error": "Failed to fetch filing"}, status_code=500
        )




@app.post("/api/upload-pdf")
async def upload_pdf(file: UploadFile = File(...)):
    if file.content_type != "application/pdf":
        return JSONResponse(status_code=400, content={"error": "Only PDF files are allowed."})

    file_bytes = await file.read()
    return await run_in_threadpool(process_pdf, file_bytes, file.filename)

def process_pdf(file_bytes: bytes, filename: str):
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    zip_stream = io.BytesIO()
    with zipfile.ZipFile(zip_stream, "w", zipfile.ZIP_DEFLATED) as zipf:
        for i, page in enumerate(doc):
            pix = page.get_pixmap(dpi=150)
            zipf.writestr(f"page_{i+1}.png", pix.tobytes("png"))

    zip_stream.seek(0)
    return StreamingResponse(
        zip_stream,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={filename.replace('.pdf', '')}_images.zip"}
    )


@app.get("/send-tasks-email")
async def send_tasks_email():
    """Fetch tasks and send them via email (hardcoded sender/recipient)."""
    try:
        data = await fetch_tasks_for_today()
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Failed to fetch tasks: {e}")

    html_body = build_html_table_grouped(data)
    subject = f"Daily Tasks Report — {date.today().isoformat()}"
    RECIPIENTS = ["megha.punjabi@decipherfinancials.com","luv.ratan@decipherfinancials.com"]

    try:
        await asyncio.to_thread(send_email, html_body, subject)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to send email: {e}")

    return {"status": "ok", "sent_to": RECIPIENTS, "count_users": len(data)}



@app.get("/filing-summary")
def get_filing_summary(url: str = Query(..., description="Base URL of the filing (without /FilingSummary.xml)")):
    """
    FastAPI endpoint to fetch and parse SEC FilingSummary.xml.
    Returns both flat reports and grouped reports by menuCategory.
    """
    try:
        # Fetch & parse FilingSummary.xml
        parsed_result = fetch_and_parse_filing_summary(url)
        
        # Group by menuCategory
        grouped = group_reports_by_category(parsed_result)
    
        # Combine both outputs
        response = {
            "groupedReports": grouped
        }

        return response

    except requests.exceptions.RequestException as e:
        raise HTTPException(status_code=502, detail=f"Error fetching FilingSummary.xml: {str(e)}")
    except ET.ParseError as e:
        raise HTTPException(status_code=500, detail=f"Error parsing XML: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Unexpected error: {str(e)}")
    

@app.get("/extract_table")
def extract_table(url: str = Query(..., description="SEC filing URL")):
    try:
        df = extract_first_table_from_sec(url, HEADERS)

        if df.empty:
            raise HTTPException(status_code=404, detail="No table data found on the page.")

        # Clean the DataFrame
        df = df.replace([float("inf"), float("-inf")], None)
        df = df.where(pd.notnull(df), None)

        return JSONResponse(content=jsonable_encoder(df.to_dict(orient="records")))

    except requests.exceptions.RequestException as e:
        raise HTTPException(status_code=400, detail=f"Request failed: {str(e)}")
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@app.get("/fetch-filing-html", response_class=Response)
async def fetch_sec_html(url: str = Query(..., description="SEC HTML file URL")):
    """
    Fetch a complete SEC filing HTML file from sec.gov and return sanitized HTML.
    """
    try:
        decoded_url = unquote(url.strip())

        if not decoded_url.startswith("https://www.sec.gov/Archives/edgar/data/"):
            raise HTTPException(status_code=400, detail="Invalid SEC URL")

        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                          "AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/121.0.0.0 Safari/537.36 (luv.ratan@decipherfinancials.com)",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
            "Referer": "https://www.sec.gov/",
        }

        async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
            response = await client.get(decoded_url, headers=headers)

        if response.status_code != 200:
            raise HTTPException(
                status_code=response.status_code,
                detail=f"SEC returned {response.status_code}: {response.text[:300]}",
            )

        # ✅ Clean and sanitize HTML before returning
        raw_html = response.text
        soup = BeautifulSoup(raw_html, "lxml")

        # Remove scripts, meta, and style elements for safety
        for tag in soup(["script", "meta", "link", "noscript", "style"]):
            tag.decompose()

        # Optional: pretty-print and compress newlines
        clean_html = " ".join(soup.prettify().split())

        # ✅ Return as actual HTML, not JSON/text
        return Response(content=clean_html, media_type="text/html")

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


#-----------------------------------------------

@app.post("/process_excel")
async def process_excel(
    file: UploadFile = File(...),
    accession_number: str = Form(...),
):
    """
    Accepts an Excel file, performs processing, and returns the processed Excel.
    """
    try:
        # 🧠 Read the uploaded Excel into memory
        file_bytes = await file.read()
        wb = load_workbook(BytesIO(file_bytes))

        # 🧹 Example: process all sheets and re-save
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                data = ws.values
                cols = next(data)
                df = pd.DataFrame(data, columns=cols)

                #Remove Blank columns
                df = df.replace(r'^\s*$', pd.NA, regex=True)  # Treat empty strings as NaN
                df = df.dropna(axis=1, how="all")
                
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        # 📤 Return processed Excel
        output.seek(0)
        headers = {
            "Content-Disposition": f"attachment; filename={accession_number}_processed.xlsx"
        }

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    except Exception as e:
        print(f"❌ process_excel error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    

# Agentic AI Toolkit ----------------------------------------------------------------------------------

# --- Utility function to read config ---
def get_bdc_config(bdc_name: str):

    json_path = os.path.join(os.path.dirname(__file__), "agentic_ai_toolkit", "cases.json")


    # Load JSON safely
    with open(json_path, "r", encoding="utf-8") as f:
        cases = json.load(f)

    # Find and return the matching case
    for case in cases:
        if case.get("BDC_name", "").strip().lower() == bdc_name.strip().lower():
            return case

    raise ValueError(f"BDC '{bdc_name}' not found in {json_path}")


# Standardizing data-----------------------------------------------------------------

@app.post("/agentic_Soi_renaming")
async def process_excel(
    file: UploadFile = File(...),
    bdc_name: str = Query(...)    # ⬅️ Required, no default
):
    try:
        # Step 1: Print bdc_name to verify input
        print(f"Received BDC name: {bdc_name}")
        # Step 1: Read uploaded Excel directly into memory
        content = await file.read()
        excel_bytes = io.BytesIO(content)

        # Step 2: Get BDC config
        config = get_bdc_config(bdc_name)
        debt_and_equity_division = config.get("debt_and_equity_division", False)

        # Step 3: Normalize everything into a list of DataFrames
        if not debt_and_equity_division:
            # append_excel_tabs returns a single DataFrame → wrap it in a list
            dataframes_list = [append_excel_tabs(excel_bytes)]
        else:
            # process_excel_file already returns a list
            dataframes_list = process_excel_file(excel_bytes)
        df_Excel = pd.concat(dataframes_list)
        df_Excel.to_excel("temp.xlsx", engine="openpyxl")

        # Step 4: Process each DataFrame individually
        all_cleaned_tables = []
        count =0
        for df in dataframes_list:
            if df is None or df.empty:
                continue

            # Each call returns a list of DataFrames → extend to collect all
            all_cleaned_tables.append(process_header_renaming(df))
            print(all_cleaned_tables[count].head())
            count+=1
        # Step 5: Concatenate all cleaned DataFrames into one
        if not all_cleaned_tables:
            raise ValueError("No valid data found after cleaning.")

        final_df = pd.concat(all_cleaned_tables, ignore_index=True)

        # ✅ Step 6: Make DataFrame JSON-safe
        final_df = final_df.replace({np.nan: None, np.inf: None, -np.inf: None})

        # Step 7: Return JSON result
        return JSONResponse(content={
            "status": "success",
            "bdc_name": bdc_name,
            "rows": int(len(final_df)),
            "data": final_df.to_dict(orient="records")
        })

    except ValueError as ve:
        raise HTTPException(status_code=404, detail=str(ve))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/post_processing_and_download_soi")
async def process_bdc_json(
    file: UploadFile = File(...),
    bdc_name: str = Form(...)
):
    """
    Endpoint to process a BDC JSON file and download the processed Excel file.
    Accepts:
      - file: JSON file to process
      - bdc_name: string argument representing the BDC name
    """
    try:
        if not file.filename.endswith(".json"):
            raise HTTPException(status_code=400, detail="Please upload a valid .json file")

        # Read and decode the uploaded file
        content = await file.read()
        json_content = json.loads(content.decode("utf-8"))
        print(bdc_name)

        # Extract data
        if isinstance(json_content, dict) and "data" in json_content:
            json_data = json_content["data"]
        elif isinstance(json_content, list):
            json_data = json_content
        else:
            raise ValueError("Invalid JSON format")

        df = pd.DataFrame(json_data)

        # ✅ Create and close temporary file manually
        tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_path = tmp_file.name
        tmp_file.close()

        if bdc_name == "Apollo Debt Solutions BDC":
             extract_first_of_contiguous_group(df, output_path=output_path)
        else:    
            # You can now use the bdc_name variable here in processing
            process_bdc(df, output_path)


        # ✅ Ensure file exists and is closed
        if not os.path.exists(output_path):
            raise HTTPException(status_code=500, detail="Excel output file not found")

        return FileResponse(
            path=output_path,
            filename=f"Cleaned_{bdc_name}_BDC_Data.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print("❌ Internal Error:", str(e))
        raise HTTPException(status_code=500, detail=f"Post Processing error: {str(e)}")


# -------------------------------------------------------------------------------------------------------------------------

#Notifications

from new_docs_notification import function_to_fetch_new_docs


@app.get("/store-notifications")
async def store_end_of_day_notifications():
    try:
        result = await function_to_fetch_new_docs("endofday")
        return result
    except Exception as e:
        logging.error(f"Failed to fetch new documents: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch new documents: {e}")


@app.get("/store-notifications-today")
async def store_today_notifications():
    try:
        result = await function_to_fetch_new_docs("today")
        return result
    except Exception as e:
        logging.error(f"Failed to fetch today's documents: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch today's documents: {e}")
    

@app.get("/reset-notifications")
async def reset_notifications():
    try:
        result = await function_to_fetch_new_docs("reset")
        return result
    except Exception as e:
        logging.error(f"Failed to fetch new documents: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch new documents: {e}")
