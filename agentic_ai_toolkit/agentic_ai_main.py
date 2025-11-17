from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import JSONResponse
import pandas as pd
import numpy as np
import openpyxl
import tempfile
import os
import json
import anthropic
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font
import re
from agentic_ai_toolkit.bifurcation_cases import bifurcate_coupon_columns
app = FastAPI(title="Excel Standardization API")



# ======== Debt and Equity ======== #


def detect_type(df: pd.DataFrame, max_rows: int = 5) -> str:
    """
    Detects if the table is one of:
    'debt', 'equity', 'stocks', 'equipment',
    'clo_equity', 'investment_funds', 'subordinate_debt',
    'money_market_funds', or 'unknown'
    based on content in first few rows.
    """

    # --- Specific patterns first ---
    subordinate_debt = r"\bsubordinated debt\b(?=\s|[-:;,%\(\)]|$)"
    clo_equity_pattern = r"\bcollateralized loan obligation(s)?\b|\bclo equity\b"
    equipment_financing_pattern = r"\bequipment financing\b"
    investment_funds_pattern = r"\binvestment fund(s)?\b"
    money_market_pattern = r"\bmoney market fund(s)?\b|\binvestment(s)? in money market\b"

    # --- Broader patterns later ---
    debt_pattern = r"\bdebt\b"
    equity_pattern = r"\bequity\b"
    stocks_pattern = r"\bstock\b"

    for i in range(min(max_rows, len(df))):
        # ✅ Normalize non-breaking spaces and lowercase text
        row_text = ' '.join([str(cell).lower().replace('\xa0', ' ') for cell in df.iloc[i].fillna('')])

        # Specific first
        if re.search(subordinate_debt, row_text):
            return "subordinate_debt"
        if re.search(clo_equity_pattern, row_text):
            return "clo_equity"
        if re.search(equipment_financing_pattern, row_text):
            return "equipment"
        if re.search(investment_funds_pattern, row_text):
            return "investment_funds"
        if re.search(money_market_pattern, row_text):
            return "money_market_funds"

        # General later
        if re.search(debt_pattern, row_text):
            return "debt"
        if re.search(equity_pattern, row_text):
            return "equity"
        if re.search(stocks_pattern, row_text):
            return "stocks"

    return "unknown"


def process_excel_file(input_path: str) -> list[pd.DataFrame]:
    """
    Reads an Excel file and groups consecutive tables of the same detected type.
    Returns a list of concatenated DataFrames (each block = one table group).
    """
    xl = pd.ExcelFile(input_path, engine="openpyxl")
    all_blocks = []

    current_type = None
    current_block = []
    misc_block = []

    for sheet_name in xl.sheet_names:
        df = xl.parse(sheet_name, header=None, dtype=str)

        # Skip empty sheets
        if df.dropna(how="all").empty:
            continue

        detected_type = detect_type(df)

        # ✅ Include the new 'money_market_funds' category here
        if detected_type in [
            "debt", "equity", "stocks", "equipment",
            "clo_equity", "investment_funds",
            "subordinate_debt", "money_market_funds"
        ]:
            if current_type is None:
                current_type = detected_type
                current_block = [df]
            elif detected_type == current_type:
                current_block.append(df)
            else:
                combined_df = pd.concat(current_block, ignore_index=True)
                all_blocks.append(combined_df)

                current_type = detected_type
                current_block = [df]

        else:
            if current_type is not None:
                current_block.append(df)
            else:
                misc_block.append(df)

    if current_block:
        combined_df = pd.concat(current_block, ignore_index=True)
        all_blocks.append(combined_df)

    if misc_block:
        combined_misc = pd.concat(misc_block, ignore_index=True)
        all_blocks.append(combined_misc)

    return all_blocks



# ======== Existing Helper Functions ======== #

def append_excel_tabs(file_like) -> pd.DataFrame:
    """
    Reads all sheets of an Excel file (file path or BytesIO),
    appends them vertically into one DataFrame, force-aligning columns.
    """

    # Read all sheets
    all_sheets = pd.read_excel(file_like, sheet_name=None, header=0)

    # Collect all DataFrames
    sheet_dfs = []

    for sheet_name, df in all_sheets.items():
        if df.empty:
            continue

        # Normalize column names
        df.columns = (
            df.columns.astype(str)
            .str.strip()
            .str.lower()
        )

        sheet_dfs.append(df)

    if not sheet_dfs:
        return pd.DataFrame()  # No valid data

    # Force append all sheets (align columns automatically)
    appended_df = pd.concat(sheet_dfs, ignore_index=True, sort=False).fillna("")
    pattern = re.compile(r"business\s*description", re.IGNORECASE)
    # Drop "Business Description" columns (case-insensitive)
    cols_to_drop = [col for col in df.columns if pattern.search(col)]
    if cols_to_drop:
        appended_df = appended_df.drop(columns=cols_to_drop, errors="ignore")

    return appended_df


def fully_normalize_df(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = (
        df.columns.astype(str)
        .str.replace(r"\s+", " ", regex=True)
        .str.replace("\xa0", " ", regex=False)
        .str.strip()
    )
    def clean_cell(x):
        if isinstance(x, str):
            return x.replace("\xa0", " ").replace("\u200b", "").strip()
        return x
    return df.map(clean_cell)


def renaming_columns_through_claude(df):
    """
    Standardize DataFrame column names using Claude AI with optimized prompting for Claude 4.5.
    
    Args:
        df: pandas DataFrame with financial data
        
    Returns:
        dict: Standardized records with mapped column names
    """
    # Initialize client
    client = anthropic.Anthropic(
        api_key="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"
    )
    
    # Handle small DataFrames safely
    head_n = min(len(df), 20)
    sample_n = min(len(df), 20)
    df_new = pd.concat(
        [df.head(head_n), df.sample(n=sample_n, replace=False)],
        ignore_index=True
    )
    
    # Convert to dict
    df_dict = df_new.to_dict(orient="list")
    
    # Standard names
    standard_columns = [
        "Portfolio Company",
        "Reported Industry",
        "Reported Investment Type",
        "Total Coupon",
        "Reference Rate",
        "Spread",
        "Floor",
        "PIK",
        "Ceiling",
        "ETP",
        "Par",
        "Cost",
        "Shares",
        "FV",
        "Percent of Net Assets",
        "Notes",
        "Origin Date",
        "Maturity",
    ]
    
    few_shot_example = """
Example:
Given the table dictionary:
{
  "Investments—non-controlled/non-affiliated (1)": [...],
  "Industry": [...],
  "Reference Rate (2)": [...],
  "Spread (2)": [...],
  "Interest Rate (2)": [...],
  "Acquisition Date": [...],
  "Maturity Date": [...],
  "Par/ Principal Amount *": [...],
  "Unnamed: 55": [...],
  "Amortized Cost\\xa0(4)": [...],
  "Unnamed: 61": [...],
  "Fair Value (5)": [...],
  "Unnamed: 67": [...],
  "% of Net\\xa0Assets": [...],
  "Unnamed: 74": [...],
  "Unnamed: 6": [...],
  "Unnamed: 12": [...]
}

Return using structure_records tool:
{
  "records": {
    "Portfolio Company": ["Investments—non-controlled/non-affiliated (1)"],
    "Reported Industry": ["Industry"],
    "Reference Rate": ["Reference Rate (2)"],
    "Spread": ["Spread (2)"],
    "Total Coupon": ["Interest Rate (2)"],
    "Origin Date": ["Acquisition Date"],
    "Maturity": ["Maturity Date"],
    "Par": ["Par/ Principal Amount *", "Unnamed: 55"],
    "Cost": ["Amortized Cost\\xa0(4)", "Unnamed: 61"],
    "FV": ["Fair Value (5)", "Unnamed: 67"],
    "Percent of Net Assets": ["% of Net\\xa0Assets", "Unnamed: 74"],
    "Notes": ["Unnamed: 6", "Unnamed: 12"],
    "Reported Investment Type": []
  }
}
"""
    
    # Optimized prompt for Claude 4.5
    prompt = f"""You are analyzing a financial data table to standardize column names.

# TASK
Map the input columns to these standard fields. Only include fields that exist in the data. Ignore columns that don't match any standard field.

<standard_columns>
{', '.join(standard_columns)}
</standard_columns>

# MAPPING RULES
Apply these rules in exact priority order (1 → 6):

**Rule 1: Total Coupon**
- Columns named "Interest Rate", "Cash Rate", "Coupon", "Coupon (%)" with fixed percentage values
- Examples: "10.81%", "9.30% (incl. 1.97% PIK)", "7.40% cash / 4.00% PIK"
- Map the ENTIRE value as-is, do not extract PIK separately

**Rule 2: Spread**
- Columns showing "Spread Above Index" or "Spread" with format like "SF + 6.50%", "P + 5.25%"
- Map the entire value as-is

**Rule 3: Reference Rate**
- Columns with ONLY benchmark names: "SOFR", "LIBOR", "SONIA", "Prime Rate"
- NOT values like "SOFR + 6.00%" (that's Spread per Rule 2)

**Rule 4: PIK**
- Dedicated PIK columns only (labeled "PIK", "Payment-in-Kind")
- Do NOT extract PIK from within other values

**Rule 5: Floor**
- Explicit Floor columns only
- Do NOT extract floors from text like "SOFR + 3.50% (floor 1.00%)"

**Rule 6: Notes**
- "Unnamed: X" columns containing only footnote markers: (1), (*), (†), etc.
- Do not extract footnotes from within other values

**Other Standard Mappings:**
- "Portfolio Company" ← company/investment names
- "Reported Industry" ← industry/sector columns
- "Reported Investment Type" ← investment type/class
- "Par" ← par/principal amount columns
- "Cost" ← amortized cost/book value columns
- "FV" ← fair value/market value columns
- "Percent of Net Assets" ← percentage of net assets
- "Origin Date" ← acquisition/origination date
- "Maturity" ← maturity date

# EXAMPLE
{few_shot_example}

# INPUT DATA
<df_dict>
{df_dict}
</df_dict>

# INSTRUCTIONS
1. Examine each input column name and its sample values
2. Apply mapping rules 1-6 in priority order
3. Group related columns (including adjacent "Unnamed: X" columns) under the same standard field
4. Only include standard fields that have matching input columns
5. For required fields (Portfolio Company, Reported Investment Type, Cost, FV) that don't exist in input, return empty arrays []
6. Use the structure_records tool to return your mapping

Be strict: only map columns that clearly match the rules. Do not hallucinate or infer fields."""

    # Enhanced function schema with descriptions
    function_schema = {
        "name": "structure_records",
        "description": "Transform raw investment data into a clean, structured format by mapping input columns to standardized field names",
        "input_schema": {
            "type": "object",
            "properties": {
                "records": {
                    "type": "object",
                    "description": "Dictionary mapping standardized field names to arrays of their corresponding input column names",
                    "properties": {
                        "Portfolio Company": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Input column(s) containing company or investment names"
                        },
                        "Reported Industry": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Input column(s) containing industry or sector classifications"
                        },
                        "Reported Investment Type": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Input column(s) containing investment types (e.g., Senior Secured Loan, Equity, Subordinated Debt)"
                        },
                        "Total Coupon": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Input column(s) containing full interest rates including PIK components"
                        },
                        "Reference Rate": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Input column(s) containing only benchmark rate names (SOFR, LIBOR, etc.)"
                        },
                        "Spread": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Input column(s) containing spread above index (e.g., 'SF + 6.50%')"
                        },
                        "Floor": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Input column(s) containing interest rate floors"
                        },
                        "PIK": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Input column(s) containing dedicated PIK (Payment-in-Kind) rates"
                        },
                        "Ceiling": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Input column(s) containing interest rate ceilings or caps"
                        },
                        "ETP": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Input column(s) containing Equity Transfer Pricing or similar"
                        },
                        "Par": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Input column(s) containing par value or principal amount"
                        },
                        "Cost": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Input column(s) containing amortized cost or book value"
                        },
                        "Shares": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Input column(s) containing share counts or units"
                        },
                        "FV": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Input column(s) containing fair value or market value"
                        },
                        "Percent of Net Assets": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Input column(s) containing percentage of net assets"
                        },
                        "Notes": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Input column(s) containing footnotes or reference markers"
                        },
                        "Origin Date": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Input column(s) containing acquisition or origination dates"
                        },
                        "Maturity": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Input column(s) containing maturity or expiration dates"
                        },
                    },
                    "required": [
                        "Portfolio Company",
                        "Reported Investment Type",
                        "Cost",
                        "FV",
                    ],
                    "additionalProperties": False
                }
            },
            "required": ["records"],
        },
    }
    
    # System instruction for Claude 4.5
    system_instruction = """You MUST use the structure_records tool to respond. Do not provide explanations or text outside the tool call.

When mapping columns:
1. Apply the semantic rules in exact priority order (1-6)
2. Only include standard fields that have matching input columns
3. For required fields that don't exist in the input, use empty arrays []
4. Group related columns (including adjacent Unnamed columns) logically based on position and content
5. Be conservative - only map columns that clearly match the rules"""
    
    # API call with optimizations for Claude 4.5
    response = client.messages.create(
        model="claude-sonnet-4-5-20250929",
        max_tokens=15000,
        temperature=0,
        system=system_instruction,  # Explicit system instruction
        tools=[function_schema],
        tool_choice={"type": "tool", "name": "structure_records"},  # Force tool usage
        messages=[{"role": "user", "content": prompt}],
    )
    
    # Extract and validate response
    response_dict = json.loads(response.model_dump_json())
    return response_dict


def format_output_records(group_list):
    raw_data = group_list["content"][0]["input"]["records"]
    if isinstance(raw_data, dict):
        return raw_data
    cleaned = raw_data.encode("utf-8", "ignore").decode("unicode_escape").replace("\xa0", " ")
    return json.loads(cleaned)


def concat_and_rename_columns(df, groups_dict):
    new_df = pd.DataFrame()
    for new_col, col_list in groups_dict.items():
        available_cols = [col for col in col_list if col in df.columns]
        if not available_cols:
            new_df[new_col] = ""
            continue
        concatenated = (
            df[available_cols]
            .astype(str)
            .apply(lambda row: " ".join(x for x in row if x and x.lower() != "nan"), axis=1)
        )
        new_df[new_col] = concatenated
    new_df.fillna("-", inplace=True)
    return new_df


def process_header_renaming(appended_df: pd.DataFrame) -> pd.DataFrame:
    """
    Processes a combined DataFrame in memory by:
    - Normalizing the data
    - Renaming columns using Claude
    - Formatting and standardizing the output
    Returns a single cleaned and standardized DataFrame.
    """

    # Step 1: Normalize DataFrame (clean text, strip spaces, fix NaNs, etc.)
    normalized_df = fully_normalize_df(appended_df)

    # Step 2: Get renaming schema using Claude (AI-based column mapping)
    group_list = renaming_columns_through_claude(normalized_df)

    # Step 3: Convert Claude’s schema output into usable format
    parsed_data = format_output_records(group_list)

    # Step 4: Apply final renaming and structure standardization
    standardized_table = concat_and_rename_columns(normalized_df, parsed_data)

    # Step 5: Return the cleaned DataFrame directly
    return standardized_table




# ======================================================
#  VERSION 1: When "Reported Industry" is missing or empty
# ======================================================

def Cb_version1(df: pd.DataFrame) -> pd.DataFrame:
    print("cb1")
    required_cols = {"Portfolio Company", "Reported Investment Type"}
    if not required_cols.issubset(df.columns):
        raise ValueError(f"Missing required columns: {required_cols}")

    # Ensure 'Reported Industry' exists and has correct dtype
    if "Reported Industry" not in df.columns:
        df["Reported Industry"] = pd.Series(dtype="object")
    else:
        df["Reported Industry"] = df["Reported Industry"].astype("object")

    # Clean blank strings -> NaN (no inplace to avoid chained-assignment warning)
    df["Reported Industry"] = (
        df["Reported Industry"]
        .replace(r'^\s*$', np.nan, regex=True)
        .infer_objects(copy=False)
    )

    # 🧩 Fill Reported Industry where Investment Type is blank
    mask_blank_investment_type = (
        df["Reported Investment Type"].isna() |
        (df["Reported Investment Type"].astype(str).str.strip() == "")
    )

    # 🧠 Ignore rows where 'Portfolio Company' contains 'total', 'debt', or 'equity'
    mask_valid_company = ~df["Portfolio Company"].astype(str).str.contains(
        r"total|debt|equity", case=False, na=False
    )

    # ✅ Final combined mask
    final_mask = mask_blank_investment_type & mask_valid_company


    # Combine both masks
    mask_valid_company = mask_blank_investment_type & final_mask

    df["Reported Industry"] = df["Reported Industry"].astype("object")
    df.loc[mask_valid_company, "Reported Industry"] = (
        df.loc[mask_valid_company, "Portfolio Company"]
        .astype("object")
        .where(lambda x: x.notna(), np.nan)
    )

    # Convert blanks or whitespace-only values to NaN before forward-fill
    df["Portfolio Company"] = df["Portfolio Company"].replace(r'^\s*$', np.nan, regex=True)
    df["Reported Industry"] = df["Reported Industry"].replace(r'^\s*$', np.nan, regex=True)

    # Forward fill
    df["Reported Industry"] = df["Reported Industry"].ffill()
    df["Portfolio Company"] = df["Portfolio Company"].ffill()

    # Drop rows where Investment Type is blank
    df = df[
        df["Reported Investment Type"].notna() &
        (df["Reported Investment Type"].astype(str).str.strip() != "")
    ].copy()

    # Clean money columns
    money_cols = ["Cost", "Par", "FV"]
    for col in money_cols:
        if col in df.columns:
            df[col] = (
                df[col]
                .astype(str)
                .str.replace("$", "", regex=False)
                .str.replace("£", "", regex=False)
                .str.replace(",", "", regex=False)
                .str.replace("(", "-", regex=False)
                .str.replace(")", "", regex=False)
                .str.replace(" ", "", regex=False)
            )
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Define final column order
    final_columns = [
        "Portfolio Company", "Non-Accrual", "Reported Industry", "DF Industry",
        "Reported Investment Type", "DF Investment Type", "Company's Stated Coupon",
        "Fixed/Float", "Floor", "Ceiling", "ETP", "Reference Rate", "Cash Spread",
        "PIK Spread", "Total Spread", "Total Coupon", "Origin Date", "Maturity",
        "FMV Measurements", "Shares", "Par", "Cost", "FV", "$ Mark", "% Mark", "Notes"
    ]

    # Add missing columns
    for col in final_columns:
        if col not in df.columns:
            df[col] = np.nan

    # Keep extra columns at the end
    extra_columns = [col for col in df.columns if col not in final_columns]
    df = df[final_columns + extra_columns]

    # Derived metrics
    df["$ Mark"] = df["FV"] - df["Cost"]
    df["% Mark"] = np.where(df["Cost"] != 0, (df["$ Mark"] / df["Cost"]) * 100, np.nan)

    # Robust date parsing
    def safe_parse_date(series: pd.Series) -> pd.Series:
        parsed = pd.to_datetime(series, errors="coerce", format="%Y-%m-%d")
        if parsed.isna().sum() > len(series) * 0.5:
            parsed = pd.to_datetime(series, errors="coerce", format="%m/%d/%Y")
        if parsed.isna().sum() > len(series) * 0.5:
            parsed = pd.to_datetime(series, errors="coerce", infer_datetime_format=True)
        return parsed


    df["Maturity"] = safe_parse_date(df["Maturity"])
    df["Origin Date"] = safe_parse_date(df["Origin Date"])

    return df

# ======================================================
#  VERSION 2: When "Reported Investment Type" is missing or empty
# ======================================================

def Cb_version2(df: pd.DataFrame) -> pd.DataFrame:
    required_cols = {"Portfolio Company", "Reported Investment Type"}
    print("cb2")
    if not required_cols.issubset(df.columns):
        raise ValueError(f"Missing required columns: {required_cols}")

    # Ensure 'Reported Industry' exists and has correct dtype
    if "Reported Industry" not in df.columns:
        df["Reported Industry"] = pd.Series(dtype="object")
    else:
        df["Reported Industry"] = df["Reported Industry"].astype("object")

    # Clean blank strings -> NaN (no inplace to avoid chained-assignment warning)
    df["Reported Industry"] = df["Reported Industry"].replace(r'^\s*$', np.nan, regex=True)

    # Fill Reported Industry where Investment Type is blank
    mask_blank_investment_type = (
        df["Reported Investment Type"].isna() |
        (df["Reported Investment Type"].astype(str).str.strip() == "")
    )

    df["Reported Industry"] = df["Reported Industry"].astype("object")
    df.loc[mask_blank_investment_type, "Reported Industry"] = (
        df.loc[mask_blank_investment_type, "Portfolio Company"]
        .astype("object")
        .where(lambda x: x.notna(), np.nan)
    )

    # Forward fill
    df["Reported Industry"] = df["Reported Industry"].ffill()
    df["Portfolio Company"] = df["Portfolio Company"].ffill()

    # Drop rows where Investment Type is blank
    df = df[
        df["Reported Investment Type"].notna() &
        (df["Reported Investment Type"].astype(str).str.strip() != "")
    ].copy()

    # Clean money columns
    money_cols = ["Cost", "Par", "FV"]
    for col in money_cols:
        if col in df.columns:
            df[col] = (
                df[col]
                .astype(str)
                .str.replace("$", "", regex=False)
                .str.replace("£", "", regex=False)
                .str.replace(",", "", regex=False)
                .str.replace("(", "-", regex=False)
                .str.replace(")", "", regex=False)
                .str.replace(" ", "", regex=False)
            )
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Define final column order
    final_columns = [
        "Portfolio Company",  "Non-Accrual", "Reported Industry", "DF Industry",
        "Reported Investment Type", "DF Investment Type", "Company's Stated Coupon", "Fixed/Float", "Floor",
        "Ceiling", "ETP", "Reference Rate", "Cash Spread", "PIK Spread", "Total Spread", "Total Coupon",
        "Origin Date", "Maturity", "FMV Measurements", "Shares", "Par", "Cost", "FV", "$ Mark", "% Mark", "Notes"
    ]

    # Add missing columns
    for col in final_columns:
        if col not in df.columns:
            df[col] = np.nan

    # Keep extra columns at the end
    extra_columns = [col for col in df.columns if col not in final_columns]
    df = df[final_columns + extra_columns]

    # Derived metrics
    df["$ Mark"] = df["FV"] - df["Cost"]
    df["% Mark"] = np.where(df["Cost"] != 0, (df["$ Mark"] / df["Cost"]) * 100, np.nan)

    # Robust date parsing
    def safe_parse_date(series: pd.Series) -> pd.Series:
        # Try standard ISO first (YYYY-MM-DD), then US-style (MM/DD/YYYY)
        parsed = pd.to_datetime(series, errors="coerce", format="%Y-%m-%d")
        if parsed.isna().sum() > len(series) * 0.5:
            parsed = pd.to_datetime(series, errors="coerce", format="%m/%d/%Y")
        return parsed

    df["Maturity"] = safe_parse_date(df["Maturity"])
    df["Origin Date"] = safe_parse_date(df["Origin Date"])

    return df


# ======================================================
#  VERSION 3: When everything is there
# ======================================================

def Cb_version3(df: pd.DataFrame) -> pd.DataFrame:
    """
    Version 3:
    Applies standard money column cleaning, derived metric calculations,
    date parsing, and column reordering when no specific filling logic applies.
    """
    print("cb3")

    # Clean money columns
    money_cols = ["Cost", "Par", "FV"]
    for col in money_cols:
        if col in df.columns:
            df[col] = (
                df[col]
                .astype(str)
                .str.replace("$", "", regex=False)
                .str.replace("£", "", regex=False)
                .str.replace(",", "", regex=False)
                .str.replace("(", "-", regex=False)
                .str.replace(")", "", regex=False)
                .str.replace(" ", "", regex=False)
            )
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Define final column order
    final_columns = [
        "Company Name", "CIK", "Ticker", "Filing Date", "DF Co. ID", "DF Asset ID", "DF Sec ID", "DF PortCom ID",
        "Portfolio Company", "Unified Asset name(DF)", "Non-Accrual", "Reported Industry", "DF Industry",
        "Reported Investment Type", "DF Investment Type", "Company's Stated Coupon", "Fixed/Float", "Floor",
        "Ceiling", "ETP", "Reference Rate", "Cash Spread", "PIK Spread", "Total Spread", "Total Coupon",
        "Origin Date", "Maturity", "FMV Measurements", "Shares", "Par", "Cost", "FV", "$ Mark", "% Mark", "Notes"
    ]

    for col in final_columns:
        if col not in df.columns:
            df[col] = np.nan

    extra_columns = [col for col in df.columns if col not in final_columns]
    df = df[final_columns + extra_columns]

    # Derived metrics
    df["$ Mark"] = df["FV"] - df["Cost"]
    df["% Mark"] = (df["$ Mark"] / df["Cost"]) * 100

    # Parse dates
    df["Maturity"] = pd.to_datetime(df["Maturity"], errors="coerce")
    df["Origin Date"] = pd.to_datetime(df["Origin Date"], errors="coerce")

    df = df[
        df["Reported Investment Type"].notna() &
        (df["Reported Investment Type"].astype(str).str.strip() != "")
    ]


    return df



def save_with_formatting(df: pd.DataFrame, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cleaned Data"
    ws.sheet_view.showGridLines = False

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
        if r_idx == 1:
            for c_idx, _ in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    wb.save(output_path)


def process_bdc(df: pd.DataFrame, output_path: str):
    """
    Automatically decides which BDC cleaning logic to apply:
    1. If 'Reported Industry' exists and is all null → run Cb_version1
    2. If 'Reported Investment Type' is missing → run Cb_version2
    3. If 'Reported Industry' does not exist at all → save as is
    """

    # Initialize cleaned_df safely
    cleaned_df = df.copy()

    # CASE 1️⃣: 'Reported Industry' exists but is all null OR not present
    if (
        ("Reported Industry" in df.columns and df["Reported Industry"].isna().all())
        or ("Reported Industry" not in df.columns)
    ):
        print("🟡 Detected: 'Reported Industry' is all null → Running Industry-filling logic (Version 1)")
        cleaned_df = Cb_version1(df)

    # CASE 2️⃣: 'Reported Investment Type' is missing entirely
    elif "Reported Investment Type" not in df.columns:
        print("🟢 Detected: 'Reported Investment Type' missing → Running Investment-Type-filling logic (Version 2)")
        cleaned_df = Cb_version2(df)

    # DEFAULT: None of the above
    else:
        print("⚪ No specific condition matched → Running Version 3 (standard cleaning)")
        cleaned_df = Cb_version3(df)

    save_with_formatting(cleaned_df, output_path)


def extract_first_of_contiguous_group(df, column_name='Twin', new_column_name='First_Twin', output_path=None):
    """
    Identifies contiguous groups of non-empty values in a column and creates
    a new column with the first item of each group.
    
    Parameters:
    -----------
    df : pd.DataFrame
        Input dataframe
    column_name : str
        Name of the column to analyze (default: 'Twin')
    new_column_name : str
        Name of the new column to create (default: 'First_Twin')
    
    Returns:
    --------
    pd.DataFrame
        Dataframe with new column added
    """
    df = df.copy()
    drop_columns_list = []

    # Identify empty rows in 'Total Coupon'
    empty_indices = df[df['Total Coupon'].isna() | (df['Total Coupon'].astype(str).str.strip() == '')].index

    # Copy values from 'Portfolio Company' to 'Twin' for those rows
    df.loc[empty_indices, 'Twin'] = df.loc[empty_indices, 'Portfolio Company']



    # Create a boolean mask for non-empty values
    # Handles None, NaN, empty strings, and whitespace-only strings
    non_empty = df[column_name].notna() & (df[column_name].astype(str).str.strip() != '')
    
    # Identify group boundaries: where value changes from empty to non-empty
    # This marks the start of each contiguous group
    group_start = non_empty & (~non_empty.shift(1, fill_value=False))
    
    # Create group IDs - each contiguous block gets a unique ID
    group_ids = group_start.cumsum()
    
    # Set group_id to 0 for empty rows (we'll ignore these)
    group_ids = group_ids.where(non_empty, 0)
    
    # Count the size of each group
    group_sizes = group_ids.map(group_ids.value_counts())
    
    # Only keep groups with 2 or more items (filter out singles)
    valid_groups = group_sizes >= 2
    
    # For each valid group ID, get the first value
    first_values = df[column_name].where(non_empty & valid_groups).groupby(group_ids).transform('first')
    
    # Create the new column
    df[new_column_name] = first_values
    df["Reported Industry"] = df["First_Twin"].ffill()
    drop_columns_list.append("First_Twin")
    df["Portfolio Company"] = df["Portfolio Company"].ffill()
    df["Twin"] = df["Twin"].ffill()
    

    # Removes rows where 'Reported Investment Type' is blank or NaN
    df = df[df['Reported Investment Type'].notna() & (df['Reported Investment Type'] != "")]
    df["Portfolio Company"] = (
    df["Portfolio Company"]
    .replace(r"^\s*$", np.nan, regex=True)  # converts "" or "   " to NaN
    .ffill())

    df["Portfolio Company"] = df["Twin"].astype(str) +"-"+ df["Portfolio Company"].astype(str)
    
    drop_columns_list.append("Twin")
    
    #Drop columns
    df = df.drop(columns=drop_columns_list)

    save_with_formatting(df, output_path)
